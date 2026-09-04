from http.server import BaseHTTPRequestHandler
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        payload = json.loads(body)

        wb = Workbook()
        wb.remove(wb.active)

        # ---- Sheet: Rekap Proyek ----
        ws_rekap = wb.create_sheet("Rekap Proyek")
        rekap_headers = payload["rekapHeaders"]
        ws_rekap.append(rekap_headers)
        for cell in ws_rekap[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        for row in payload["rekapRows"]:
            ws_rekap.append(row)

        # Auto-adjust column widths for Rekap
        for col in ws_rekap.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_rekap.column_dimensions[column].width = adjusted_width

        # ---- Sheet: Pivot Dept x Status ----
        ws_dept = wb.create_sheet("Pivot Dept x Status")
        dept_headers = ["Departemen"] + payload["allStatuses"] + ["TOTAL"]
        ws_dept.append(dept_headers)
        for cell in ws_dept[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        for dept_row in payload["deptStatusMatrix"]:
            ws_dept.append([dept_row["dept"]] + dept_row["counts"] + [dept_row["total"]])
        total_row_idx = len(payload["deptStatusMatrix"]) + 2
        ws_dept.append(["TOTAL"] + payload["deptStatusColumnTotals"] + [payload["deptStatusGrandTotal"]])
        for cell in ws_dept[total_row_idx]:
            cell.font = Font(bold=True)

        # Bar chart
        data_last_row = total_row_idx - 1
        max_col = len(dept_headers)
        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Jumlah Proyek per Departemen & Status"
        chart1.y_axis.title = "Jumlah Proyek"
        chart1.style = 10
        chart1.visible_cells_only = False
        data = Reference(ws_dept, min_col=2, max_col=max_col - 1, min_row=1, max_row=data_last_row)
        cats = Reference(ws_dept, min_col=1, max_col=1, min_row=2, max_row=data_last_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height, chart1.width = 8, 16
        ws_dept.add_chart(chart1, f"A{total_row_idx + 3}")

        # ---- Sheet: Pivot Bulan x Status ----
        ws_bulan = wb.create_sheet("Pivot Bulan x Status")
        bulan_headers = ["Bulan"] + payload["allStatuses"] + ["TOTAL"]
        ws_bulan.append(bulan_headers)
        for cell in ws_bulan[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        for month_row in payload["monthStatusMatrix"]:
            ws_bulan.append([month_row["month"]] + month_row["counts"] + [month_row["total"]])
        total_row_idx2 = len(payload["monthStatusMatrix"]) + 2
        ws_bulan.append(["TOTAL"] + payload["monthStatusColumnTotals"] + [payload["monthStatusGrandTotal"]])
        for cell in ws_bulan[total_row_idx2]:
            cell.font = Font(bold=True)

        data_last_row2 = total_row_idx2 - 1
        chart2 = LineChart()
        chart2.title = "Tren Volume Proyek per Bulan"
        chart2.y_axis.title = "Jumlah Proyek"
        chart2.style = 12
        chart2.visible_cells_only = False
        data2 = Reference(ws_bulan, min_col=2, max_col=len(bulan_headers) - 1, min_row=1, max_row=data_last_row2)
        cats2 = Reference(ws_bulan, min_col=1, max_col=1, min_row=2, max_row=data_last_row2)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        for s in chart2.series:
            s.marker.symbol = "circle"
        chart2.height, chart2.width = 8, 16
        ws_bulan.add_chart(chart2, f"A{total_row_idx2 + 3}")

        # ---- Sheet: Statistik (+ 2 pie chart) ----
        ws_stat = wb.create_sheet("Statistik")
        ws_stat.append(["Metrik", "Nilai"])
        for cell in ws_stat[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
        for r in payload["statistikRows"]:
            ws_stat.append(r)

        status_labels = payload["statusDistribution"]["labels"]
        status_values = payload["statusDistribution"]["values"]
        if status_labels:
            pie1 = PieChart()
            pie1.title = "Distribusi Status Proyek"
            # PENTING: default Excel tidak mem-plot data dari kolom/baris yang disembunyikan.
            # Karena kolom bantu E/F di bawah ini sengaja disembunyikan (hidden=True),
            # plotVisOnly HARUS di-set False, kalau tidak chart akan tampil KOSONG.
            pie1.visible_cells_only = False
            tmp_row = ws_stat.max_row + 2
            for i, (lbl, val) in enumerate(zip(status_labels, status_values)):
                ws_stat.cell(row=tmp_row + i, column=5, value=lbl)
                ws_stat.cell(row=tmp_row + i, column=6, value=val)
            data_p1 = Reference(ws_stat, min_col=6, min_row=tmp_row, max_row=tmp_row + len(status_labels) - 1)
            cats_p1 = Reference(ws_stat, min_col=5, min_row=tmp_row, max_row=tmp_row + len(status_labels) - 1)
            pie1.add_data(data_p1, titles_from_data=False)
            pie1.set_categories(cats_p1)
            pie1.dataLabels = DataLabelList()
            pie1.dataLabels.showPercent = True
            pie1.height, pie1.width = 7, 10
            ws_stat.add_chart(pie1, "D2")
            ws_stat.column_dimensions["E"].hidden = True
            ws_stat.column_dimensions["F"].hidden = True

        dept_labels = payload["deptDistribution"]["labels"]
        dept_values = payload["deptDistribution"]["values"]
        if dept_labels:
            pie2 = PieChart()
            pie2.title = "Distribusi Proyek per Departemen"
            # Sama seperti pie1 — wajib False karena kolom E/F disembunyikan juga di bagian ini.
            pie2.visible_cells_only = False
            tmp_row2 = tmp_row + len(status_labels) + 3 if status_labels else ws_stat.max_row + 2
            for i, (lbl, val) in enumerate(zip(dept_labels, dept_values)):
                ws_stat.cell(row=tmp_row2 + i, column=5, value=lbl)
                ws_stat.cell(row=tmp_row2 + i, column=6, value=val)
            data_p2 = Reference(ws_stat, min_col=6, min_row=tmp_row2, max_row=tmp_row2 + len(dept_labels) - 1)
            cats_p2 = Reference(ws_stat, min_col=5, min_row=tmp_row2, max_row=tmp_row2 + len(dept_labels) - 1)
            pie2.add_data(data_p2, titles_from_data=False)
            pie2.set_categories(cats_p2)
            pie2.dataLabels = DataLabelList()
            pie2.dataLabels.showPercent = True
            pie2.height, pie2.width = 7, 10
            ws_stat.add_chart(pie2, "D18")

        # ---- Sheet: Kesimpulan Analitik ----
        ws_kesimpulan = wb.create_sheet("Kesimpulan Analitik")
        ws_kesimpulan.cell(row=1, column=1, value="KESIMPULAN & ANALISIS NARATIF OTOMATIS")
        ws_kesimpulan["A1"].font = Font(bold=True, size=13)
        for i, line in enumerate(payload["kesimpulanLines"], start=3):
            cell = ws_kesimpulan.cell(row=i, column=1, value=line)
            cell.alignment = Alignment(wrap_text=True)
        ws_kesimpulan.column_dimensions["A"].width = 100

        # ---- Send response ----
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{payload.get("filename", "Rekap-Kaizen.xlsx")}"')
        self.end_headers()
        self.wfile.write(buffer.read())
        return